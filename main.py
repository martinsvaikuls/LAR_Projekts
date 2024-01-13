##datu manipulāciju bibliotēkas
import selenium

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service

from pywinauto.application import Application

import time

##datu uzgalabašana
import unicodedata
import csv
from openpyxl import Workbook, load_workbook

##svarīgu datu uzglabāšana
import userInformation as usrI

##datu iegūšana no tīmekļa vietnes
#"""
usrName = usrI.username
passW = usrI.password

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

url = usrI.userURL
driver.get(url)
time.sleep(1)

#ievada lietotaj vardu
find = driver.find_element(By.ID, "IDToken1")
find.send_keys(usrName)

#ievada paroli
find = driver.find_element(By.ID, "IDToken2")
find.send_keys(passW)
#input()
#"Pieteikties" -> Enter
find.send_keys('\ue007')
time.sleep(5)
#input()
find = driver.find_element(By.CLASS_NAME, "gotocal")
find.click()
time.sleep(5)
workNames = []
classTime = []
className = []

webElement = driver.find_elements(By.CLASS_NAME, "name.d-inline-block")
for i in range(len(webElement)):
    #print(webElement[i])
    workNames.append(webElement[i].text)

webElement = driver.find_elements(By.XPATH, "//div[@class='row']/div[@class='col-11']")
for i in range(len(webElement)):
    classTime.append(webElement[i].text)

webElement = driver.find_elements(By.XPATH, "//div[@class='row mt-1']/div[@class='col-11']/a")
for i in range(len(webElement)):
    className.append(webElement[i].text)
#"""
###

######
#workNames = usrI.testWorks
#className = usrI.testClasses
#classTime = usrI.testTimes
######

filePath = usrI.filePathXlsx
wb = load_workbook(filePath)
ws = wb.active
maxRows = ws.max_row
oldWorkNames = []
oldClassName = []
oldClassTime = []

##iegūt jau esošos datus 
for i in range(2,maxRows+1):
    oldWorkNames.append(str(ws['B'+ str(i)].value)) 
    oldClassName.append(str(ws['C'+ str(i)].value))
    oldClassTime.append(str(ws['D'+ str(i)].value))
    print("a")
###
print(oldClassName, oldClassTime, oldWorkNames)

##uzzināt kuri dati ir jauni un kuri vairs nepastāv
#jauni dati
newName = []
newClass = []
newTime = []

notifyOldName = []
notifyOldClass = []
notifyOldTime = []

if len(workNames) > len(oldWorkNames):
    for i in range(len(workNames)):

        if workNames[i] in oldWorkNames:
            notifyOldName.append(workNames[i])
            notifyOldClass.append(className[i])
            notifyOldTime.append(classTime[i])
        else:
            newName.append(workNames[i])
            newClass.append(className[i])
            newTime.append(classTime[i])
else: 
    for i in range(len(oldWorkNames)):

        if oldWorkNames[i] in workNames:
            notifyOldName.append(oldWorkNames[i])
            notifyOldClass.append(oldClassName[i])
            notifyOldTime.append(oldClassTime[i])
        else:
            newName.append(oldWorkNames[i])
            newClass.append(oldClassName[i])
            newTime.append(oldClassTime[i])



##ievietot datus excelī
dataAmount = len(workNames)
print("len of many", dataAmount)

for i in range(dataAmount):
    try:
        ws['B'+ str(i+2)] = str(workNames[i])
        ws['C'+ str(i+2)] = str(className[i])
        ws['D'+ str(i+2)] = str(classTime[i])
        #cellref=worksheet.cell(row=i, column=4) cellref.value=
    except:
        print("oops with the xlsx")

#wb2 = Workbook()
wb.save(filePath)
#'D:\VSCODE_Programmes\Projects\Lists\prr\data2.xlsx'
wb.close()

driver.close()
###another program
import pywinauto
from pywinauto import win32functions
import win32api

from pywinauto import keyboard, mouse


##pārbaudīt vai programma jau ir atvērta
import pyautogui

app = "Telegram"
if app not in pyautogui.getAllWindows():
    app = Application(backend = "uia").start(r"D:\Programmas\Telegram\Telegram Desktop\Telegram.exe")

######
time.sleep(8)
pywinauto.mouse.click(coords = (231, 187))
time.sleep(1)
pywinauto.mouse.click(coords = (641, 1007))
pywinauto.keyboard.send_keys("jaunie darbi")
pywinauto.keyboard.send_keys('+{ENTER}')
##!! TypeError: 'str' object cannot be interpreted as an integer   at 138+139 if in same line
for i in range(len(newName)): 
    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}") 
    helperList = newName[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')

    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}")
    helperList = newClass[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')

    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}")
    helperList = newTime[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')
    pywinauto.keyboard.send_keys('+{ENTER}')
    

pywinauto.keyboard.send_keys('+{ENTER}')
pywinauto.keyboard.send_keys("vecie darbi")
pywinauto.keyboard.send_keys('+{ENTER}')

for i in range(len(notifyOldName)):
    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}")
    helperList = notifyOldName[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')

    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}")
    helperList = notifyOldClass[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')

    for j in range(4):
        pywinauto.keyboard.send_keys("{SPACE}")
    helperList = notifyOldTime[i].split()
    for j in range(len(helperList)):
        pywinauto.keyboard.send_keys(helperList[j])
        pywinauto.keyboard.send_keys("{SPACE}")
    pywinauto.keyboard.send_keys('+{ENTER}')
    pywinauto.keyboard.send_keys('+{ENTER}')

pywinauto.keyboard.send_keys('{ENTER}')
######

time.sleep(1)




##izveidot vienu texta līniju nevis lietot trīs masīvus
"""
homeworks = []
time.sleep(1)
for data in range(workNames):
    homeworks.append(workNames[data]+','+className[data]+','+classTime[data])
"""
###

##pārveidot jauniegūtos datus, lai tos varētu ievietot csv
"""
for i in range(len(workNames)):
    workNames[i] = unicodedata.normalize("NFKD", workNames[i]).encode("ascii", "ignore")
    className[i] = unicodedata.normalize("NFKD", className[i]).encode("ascii", "ignore")
    classTime[i] = unicodedata.normalize("NFKD", classTime[i]).encode("ascii", "ignore")
"""
###

##! ievietot datus csv failā
"""
filepathcsv = usrI.filePathcsv
with open(, 'w', newline='') as csvF:
    file = csv.writer(csvF, delimiter="$")
    for i in range(len(homeworks)):
        print("doeshappen")
        try:
            file.writerow(homeworks[i])
        except:
            try:
                file.writerow(unicodedata.normalize('NFKD',homeworks[i]).encode("ascii", 'ignore'))
            except:
                print("noneOfThemWorkError")
        else:
            print("text was encoded")

csvF.close()

"""
###!
"""
for texts in classTime:
    print(texts.text)

for texts in className:
    print(texts.text)

for texts in workNames:
    print(texts.text)
"""


#input()


